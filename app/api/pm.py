"""PM Agent API endpoints.

Storage-backed endpoints (MySQL or JSON file fallback):
  POST /api/pm/chat        Phase 1: requirement mining dialog (SSE)
  POST /api/pm/generate    Phase 2: PRD generation (SSE)
  POST /api/pm/continue    Phase 2: incremental next section (SSE)
  POST /api/pm/agent       Convenience wrapper: auto-routes (SSE)
  GET  /api/pm/sessions    List sessions (persistent)
  GET  /api/pm/history/{session_id}  Chat messages
  DELETE /api/pm/sessions/{session_id}   Delete session
  GET  /api/pm/profile/{session_id}  View requirement profile
  GET  /api/pm/prd/{session_id}  View generated PRD
  GET  /api/pm/version     App info

  P1 — Data services:
  GET  /api/pm/dashboard/overview
  GET  /api/pm/dashboard/trend
  GET  /api/pm/export/sessions  (CSV or XLSX)
  GET  /api/pm/export/prds      (XLSX)
"""

import csv
import io
import json

from fastapi import APIRouter, Depends, HTTPException, Query, Request, File, UploadFile
from fastapi.responses import StreamingResponse
from loguru import logger
from sse_starlette.sse import EventSourceResponse

from app.core.auth import get_current_user, require_prd_generator
from app.models.pm import AgentRequest, ChatRequest, ContinueRequest, ExtractProposalRequest, GenerateFromProposalsRequest, GenerateRequest
# Deferred lazy accessor to avoid circular import:
# app.main → app.api.pm → (would) → app.main
def _svc():
    from app.main import get_pm_agent_service
    s = get_pm_agent_service()
    if s is None:
        raise RuntimeError("PM Agent service not initialized (server still starting)")
    return s

router = APIRouter()


async def _check_session_ownership(
    current_user: dict, session_id: str
) -> dict:
    """Check if current user owns the session (admin can access any)."""
    session = await _svc().get_session(session_id)
    if session is None:
        raise HTTPException(status_code=404, detail="Session not found")
    if current_user.get("role") != "admin" and session.get("user_id") != current_user["id"]:
        raise HTTPException(status_code=403, detail="无权访问此会话")
    return session


# ── Session management ──


@router.get("/pm/sessions")
async def pm_list_sessions(
    current_user: dict = Depends(get_current_user),
    all: bool = Query(default=False),
    user_id: str | None = Query(default=None),
    status: str | None = Query(default=None),
    limit: int = Query(default=50, le=200),
    offset: int = Query(default=0, ge=0),
):
    """List sessions with optional filters, ordered by updated_at desc.

    Regular users see only their own sessions.  Admin can use ?all=true
    to see all sessions, or pass a specific user_id.
    """
    if all and current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="仅管理员可查看所有会话")
    role = current_user.get("role")
    if all:
        uid = None  # no filter
    elif role in ("analyst", "admin"):
        uid = None  # analyst/admin 看所有会话（工作空间共享）
    elif user_id:
        uid = user_id
    else:
        uid = current_user["id"]

    sessions = await _svc().list_sessions()
    # Client-side filtering for file-fallback compat
    if uid:
        sessions = [s for s in sessions if s.get("user_id") == uid]
    if status:
        sessions = [s for s in sessions if s.get("status") == status]
    return {"success": True, "sessions": sessions[offset : offset + limit], "total": len(sessions)}


@router.get("/pm/history/{session_id}")
async def pm_get_history(
    session_id: str,
    current_user: dict = Depends(get_current_user),
    limit: int = Query(default=200, le=1000),
    offset: int = Query(default=0, ge=0),
):
    """Get chat messages for a session (persistent)."""
    await _check_session_ownership(current_user, session_id)
    messages = await _svc().get_message_history(session_id)
    return {
        "success": True,
        "session_id": session_id,
        "messages": messages[offset : offset + limit],
        "total": len(messages),
    }


@router.delete("/pm/sessions/{session_id}")
async def pm_delete_session(
    session_id: str,
    current_user: dict = Depends(get_current_user),
):
    """Delete a session and all related data."""
    await _check_session_ownership(current_user, session_id)
    deleted = await _svc().delete_session(session_id)
    await _svc().log_audit("delete", user_id=current_user["id"], session_id=session_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Session not found")
    return {"success": True, "session_id": session_id}


# ── Import / Upload ──


@router.post("/pm/import")
async def pm_import(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
):
    """Upload a .md file → create new session → AI analysis (SSE).

    Accepts multipart/form-data with a single 'file' field.
    Returns SSE events: import_analysis, content, sufficiency, import_complete.
    """
    content = await file.read()
    filename = file.filename or "untitled.md"

    logger.info(f"[import] {filename} ({len(content)} bytes)")

    async def gen():
        async for event in _svc().import_document(content, filename, current_user["id"]):
            yield {"event": "message", "data": json.dumps(event, ensure_ascii=False)}

    return EventSourceResponse(gen())


@router.post("/pm/sessions/{session_id}/upload")
async def pm_upload_session_file(
    session_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
):
    """Upload a .md file to an existing session (SSE).

    File content is injected as context for the AI to reference
    in the ongoing conversation.
    """
    content = await file.read()
    filename = file.filename or "untitled.md"

    logger.info(f"[upload] {session_id} <- {filename} ({len(content)} bytes)")
    await _check_session_ownership(current_user, session_id)

    async def gen():
        async for event in _svc().upload_to_session(session_id, content, filename):
            yield {"event": "message", "data": json.dumps(event, ensure_ascii=False)}

    return EventSourceResponse(gen())


# ── Phase 1: Chat ──


@router.post("/pm/chat")
async def pm_chat(
    request: ChatRequest,
    current_user: dict = Depends(get_current_user),
):
    """Phase 1: Requirement mining dialog (SSE)."""
    logger.info(f"[{request.session_id}] Chat: {request.message[:100]}...")

    async def gen():
        async for event in _svc().chat(
            request.message,
            request.session_id,
            user_id=current_user["id"],
            workspace_id=request.workspace_id or None,
            use_knowledge=request.use_knowledge,
            referenced_files=request.referenced_files,
        ):
            yield {"event": "message", "data": json.dumps(event, ensure_ascii=False)}

    return EventSourceResponse(gen())


# ── Phase 2: PRD Generation ──


@router.post("/pm/generate")
async def pm_generate(
    request: GenerateRequest,
    current_user: dict = Depends(require_prd_generator),
):
    """Phase 2: Trigger PRD generation (SSE)."""
    logger.info(f"[{request.session_id}] Generate PRD, mode={request.mode}")
    await _check_session_ownership(current_user, request.session_id)

    async def gen():
        async for event in _svc().generate_prd(request.session_id, request.mode):
            yield {"event": "message", "data": json.dumps(event, ensure_ascii=False)}

    return EventSourceResponse(gen())


@router.post("/pm/continue")
async def pm_continue(
    request: ContinueRequest,
    current_user: dict = Depends(require_prd_generator),
):
    """Phase 2: Continue incremental generation (SSE)."""
    logger.info(f"[{request.session_id}] Continue generation")
    await _check_session_ownership(current_user, request.session_id)

    async def gen():
        async for event in _svc().continue_generation(request.session_id):
            yield {"event": "message", "data": json.dumps(event, ensure_ascii=False)}

    return EventSourceResponse(gen())


# ── Convenience wrapper ──


@router.post("/pm/agent")
async def pm_agent(
    request: AgentRequest,
    current_user: dict = Depends(get_current_user),
):
    """Convenience wrapper: auto-routes chat/generate/continue (SSE)."""
    logger.info(f"[{request.session_id}] Agent: {request.message[:100]}...")

    async def gen():
        async for event in _svc().handle(
            request.message,
            request.session_id,
            request.mode,
            user_id=current_user["id"],
            workspace_id=request.workspace_id or None,
            use_knowledge=request.use_knowledge,
            referenced_files=request.referenced_files,
        ):
            yield {"event": "message", "data": json.dumps(event, ensure_ascii=False)}

    return EventSourceResponse(gen())


# ── Profile / PRD access ──


@router.get("/pm/profile/{session_id}")
async def pm_get_profile(
    session_id: str,
    current_user: dict = Depends(get_current_user),
):
    """Get current requirement profile."""
    await _check_session_ownership(current_user, session_id)
    profile = await _svc().get_profile(session_id)
    return {"success": True, "session_id": session_id, "profile": profile}


@router.get("/pm/prd/{session_id}")
async def pm_get_prd(
    session_id: str,
    current_user: dict = Depends(get_current_user),
):
    """Get generated PRD for a session."""
    await _check_session_ownership(current_user, session_id)
    prd = await _svc().get_prd(session_id)
    if prd is None:
        raise HTTPException(status_code=404, detail="No PRD generated for this session yet")
    return {"success": True, "session_id": session_id, "prd": prd}


# ── Dashboard (P1) ──


@router.get("/pm/dashboard/overview")
async def pm_dashboard_overview(
    current_user: dict = Depends(get_current_user),
):
    """Get aggregated dashboard overview: sessions by status, avg sufficiency, totals."""
    overview = await _svc().get_dashboard_overview()
    return {"success": True, "data": overview}


@router.get("/pm/dashboard/trend")
async def pm_dashboard_trend(
    current_user: dict = Depends(get_current_user),
    granularity: str = Query(default="day", pattern="^(day|week|month)$"),
    days: int = Query(default=30, ge=1, le=365),
):
    """Get time-series trend data of session and PRD creation counts."""
    trend = await _svc().get_trend_data(granularity, days)
    return {"success": True, "data": trend}


# ── Export (P1) ──


@router.get("/pm/export/sessions")
async def pm_export_sessions(
    current_user: dict = Depends(get_current_user),
    format: str = Query(default="csv", pattern="^(csv|xlsx)$"),
):
    """Export session list as CSV or XLSX."""
    sessions = await _svc().export_sessions()

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["session_id", "user_id", "project_name", "status",
                         "sufficiency_score", "created_at", "updated_at"])
        for s in sessions:
            writer.writerow([
                s.get("session_id", ""),
                s.get("user_id", ""),
                s.get("project_name", ""),
                s.get("status", ""),
                s.get("sufficiency_score", 0),
                s.get("created_at", ""),
                s.get("updated_at", ""),
            ])
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=sessions.csv"},
        )

    # XLSX export
    return await _export_xlsx(
        filename="sessions.xlsx",
        headers=["session_id", "user_id", "project_name", "status",
                 "sufficiency_score", "created_at", "updated_at"],
        rows=[[
            s.get("session_id", ""),
            s.get("user_id", ""),
            s.get("project_name", ""),
            s.get("status", ""),
            s.get("sufficiency_score", 0),
            s.get("created_at", ""),
            s.get("updated_at", ""),
        ] for s in sessions],
    )


@router.get("/pm/export/prds")
async def pm_export_prds(
    current_user: dict = Depends(get_current_user),
    format: str = Query(default="xlsx", pattern="^(csv|xlsx)$"),
):
    """Export generated PRDs as XLSX or CSV."""
    prds = await _svc().export_prds()

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["session_id", "project_name", "version", "mode",
                         "created_at", "markdown_preview"])
        for p in prds:
            writer.writerow([
                p.get("session_id", ""),
                p.get("project_name", ""),
                p.get("version", 1),
                p.get("mode", ""),
                p.get("created_at", ""),
                p.get("markdown_preview", ""),
            ])
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=prds.csv"},
        )

    return await _export_xlsx(
        filename="prds.xlsx",
        headers=["session_id", "project_name", "version", "mode",
                 "created_at", "markdown_preview"],
        rows=[[
            p.get("session_id", ""),
            p.get("project_name", ""),
            p.get("version", 1),
            p.get("mode", ""),
            p.get("created_at", ""),
            p.get("markdown_preview", ""),
        ] for p in prds],
    )


async def _export_xlsx(
    filename: str,
    headers: list[str],
    rows: list[list],
) -> StreamingResponse:
    """Helper to generate XLSX via openpyxl and return StreamingResponse."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Export"

    # Header row
    header_font = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font

    # Data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    for col_idx, h in enumerate(headers, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(
            len(str(h)) + 2,
            max((len(str(row[col_idx - 1])) if col_idx <= len(row) else 0)
                for row in rows) + 2 if rows else 10,
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── Version ──


@router.get("/pm/version")
async def pm_version():
    """Get PM Agent version info."""
    from app.config import config
    return {
        "app_name": config.app_name,
        "version": config.app_version,
        "llm_model": config.llm_model,
        "llm_base_url": config.llm_base_url,
    }


# ── PRD by ID ──


@router.get("/pm/prd/by-id/{prd_id}")
async def pm_get_prd_by_id(
    prd_id: str,
    current_user: dict = Depends(get_current_user),
):
    """Get a PRD by its primary key ID."""
    from app.main import get_datastore
    ds = get_datastore()
    if ds is None:
        raise HTTPException(status_code=500, detail="DataStore not initialized")
    prd = await ds.get_prd_by_id(prd_id)
    if prd is None:
        raise HTTPException(status_code=404, detail="PRD not found")
    return {"success": True, "prd": prd}


# ── Proposal Extraction ──


@router.post("/pm/extract-proposal")
async def pm_extract_proposal(
    request: ExtractProposalRequest,
    current_user: dict = Depends(get_current_user),
):
    """Extract a requirement proposal from a session and assess priority (SSE)."""
    logger.info(f"[extract-proposal] session={request.session_id}, ws={request.workspace_id}")

    # Validate session ownership
    await _check_session_ownership(current_user, request.session_id)

    async def gen():
        from app.agent.pm.proposal_extractor import extract_proposal_from_session
        from app.agent.pm.proposal_priority import assess_priority
        from app.main import get_datastore

        ds = get_datastore()
        if ds is None:
            raise RuntimeError("DataStore not initialized")

        # Step 1: Extract proposal fields from conversation
        proposal_data = None
        async for event in extract_proposal_from_session(request.session_id, ds):
            if event["type"] == "proposal_field":
                yield {
                    "event": "message",
                    "data": json.dumps(event, ensure_ascii=False),
                }
            elif event["type"] == "proposal_done":
                proposal_data = event["data"]

        if proposal_data is None:
            yield {
                "event": "message",
                "data": json.dumps({
                    "type": "proposal_done",
                    "data": {"title": "", "background": "", "pain_points": [],
                             "desired_outcome": "", "scope_note": "", "tags": []},
                }, ensure_ascii=False),
            }
            return

        # Step 2: Assess priority (emit a status event first for UI feedback)
        yield {
            "event": "message",
            "data": json.dumps({
                "type": "proposal_field",
                "field": "priority",
                "display_name": "优先级评估",
                "content": "评估中...",
            }, ensure_ascii=False),
        }
        priority, reasoning = await assess_priority(proposal_data)
        proposal_data["priority"] = priority
        proposal_data["ai_assessment"] = reasoning

        # Step 3: Get session info for submitter
        session_info = await ds.get_session(request.session_id)
        submitter_id = session_info.get("user_id", "") if session_info else ""
        if not submitter_id:
            submitter_id = current_user.get("id", "")

        # Step 4: Persist proposal to workspace
        try:
            saved = await ds.create_proposal(
                workspace_id=request.workspace_id,
                title=proposal_data.get("title", ""),
                background=proposal_data.get("background", ""),
                pain_points=proposal_data.get("pain_points", []),
                desired_outcome=proposal_data.get("desired_outcome", ""),
                scope_note=proposal_data.get("scope_note", ""),
                urgency=priority_to_urgency(priority),
                priority=priority,
                tags=proposal_data.get("tags", []),
                source_session_id=request.session_id,
                submitter_id=submitter_id,
                status="pending_review",
            )

            # Emit final event with saved proposal
            yield {
                "event": "message",
                "data": json.dumps({
                    "type": "proposal_done",
                    "data": saved,
                }, ensure_ascii=False),
            }

            logger.info(f"[extract-proposal] saved proposal={saved['id']} priority={priority}")
        except Exception as e:
            logger.error(f"[extract-proposal] save failed: {e}")
            yield {
                "event": "message",
                "data": json.dumps({
                    "type": "proposal_done",
                    "data": proposal_data,
                }, ensure_ascii=False),
            }

    return EventSourceResponse(gen())


def priority_to_urgency(priority: str) -> str:
    """Map priority level to urgency label."""
    mapping = {"P0": "critical", "P1": "high", "P2": "medium", "P3": "low"}
    return mapping.get(priority, "medium")


# ── Generate from Proposals ──


@router.post("/pm/generate-from-proposals")
async def pm_generate_from_proposals(
    request: GenerateFromProposalsRequest,
    current_user: dict = Depends(require_prd_generator),
):
    """Generate a PRD from approved proposals (SSE streaming)."""
    logger.info(
        f"[gen-from-proposals] ws={request.workspace_id}, "
        f"proposals={request.proposal_ids}"
    )

    if not request.proposal_ids:
        raise HTTPException(status_code=400, detail="No proposal IDs provided")

    async def gen():
        from app.main import get_datastore
        from app.agent.pm.phase2.planner import prd_planner
        from app.agent.pm.phase2.assembler import prd_assembler

        ds = get_datastore()
        if ds is None:
            yield {
                "event": "error",
                "data": json.dumps({"detail": "DataStore not initialized"}, ensure_ascii=False),
            }
            return

        # 1. Fetch and validate proposals
        proposals = []
        for pid in request.proposal_ids:
            p = await ds.get_proposal(pid)
            if p is None:
                yield {
                    "event": "error",
                    "data": json.dumps({"detail": f"Proposal not found: {pid}"}, ensure_ascii=False),
                }
                return
            if p.get("status") != "approved":
                yield {
                    "event": "error",
                    "data": json.dumps(
                        {"detail": f"Proposal {pid} is not approved (status: {p.get('status')})"},
                        ensure_ascii=False,
                    ),
                }
                return
            proposals.append(p)

        # 2. Merge proposals into a profile-like dict
        titles = [p.get("title", "") for p in proposals if p.get("title")]
        backgrounds = [p.get("background", "") for p in proposals if p.get("background")]
        pain_points = []
        for p in proposals:
            pts = p.get("pain_points", [])
            if isinstance(pts, list):
                pain_points.extend(pts)

        outcomes = [p.get("desired_outcome", "") for p in proposals if p.get("desired_outcome")]
        scope_notes = [p.get("scope_note", "") for p in proposals if p.get("scope_note")]

        project_name = titles[0][:80] if titles else "Combined Proposals PRD"
        elevator_pitch = " / ".join(titles[:3]) if titles else ""

        merged_profile = {
            "project_name": project_name,
            "elevator_pitch": elevator_pitch,
            "background": "\n\n".join(backgrounds),
            "pain_points": "\n".join(f"- {pt}" for pt in pain_points),
            "desired_outcome": "\n".join(outcomes),
            "scope_note": "\n".join(scope_notes),
            "business_type": "",
            "target_users": "项目干系人",
            "user_count": 0,
            "user_scale": "department",
            "urgency": "medium",
            "constraints": "",
            "custom_prompts": [],
            "source_proposal_ids": request.proposal_ids,
        }

        # 3. Plan sections
        sections = prd_planner.plan(merged_profile, mode="one_shot")
        yield {
            "event": "message",
            "data": json.dumps({
                "type": "prd_plan",
                "data": {
                    "sections": [{"key": s["key"], "title": s["title"]} for s in sections],
                    "mode": "one_shot",
                    "source": "proposals",
                },
            }, ensure_ascii=False),
        }

        # 4. Assemble PRD via assembler
        thread_id = request.session_id or f"proposals-{request.proposal_ids[0][:8]}"
        full_markdown = ""

        async for event in prd_assembler.assemble(sections, merged_profile, mode="one_shot"):
            if event.get("type") == "prd_complete":
                markdown = event["data"]["markdown"]
                full_markdown = markdown

                # 5. Save PRD
                try:
                    prd_data = {
                        "project_name": project_name,
                        "mode": "one_shot",
                        "sections": [
                            {"key": s["key"], "title": s["title"],
                             "status": s.get("status", "done")}
                            for s in sections
                        ],
                        "markdown": markdown,
                    }
                    await ds.save_prd(
                        thread_id,
                        project_name=project_name,
                        mode="one_shot",
                        sections=prd_data["sections"],
                        markdown=markdown,
                    )

                    # NOTE: source_proposal_ids will be set when save_prd
                    # is enhanced to accept the field (DataStore update upcoming)

                    full_markdown = markdown
                    logger.info(
                        f"[gen-from-proposals] PRD saved: {project_name[:60]}, "
                        f"sources={request.proposal_ids}"
                    )

                    yield {
                        "event": "message",
                        "data": json.dumps({
                            "type": "prd_complete",
                            "data": {
                                "markdown": markdown,
                                "project_name": project_name,
                                "source_proposal_ids": request.proposal_ids,
                            },
                        }, ensure_ascii=False),
                    }
                except Exception as e:
                    logger.error(f"[gen-from-proposals] Save failed: {e}")
                    yield {
                        "event": "message",
                        "data": json.dumps({
                            "type": "prd_complete",
                            "data": {"markdown": markdown, "project_name": project_name},
                        }, ensure_ascii=False),
                    }
            else:
                yield {
                    "event": "message",
                    "data": json.dumps(event, ensure_ascii=False),
                }

    return EventSourceResponse(gen())
